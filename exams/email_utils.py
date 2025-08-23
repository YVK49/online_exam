import io
from openpyxl import Workbook
from django.core.mail import EmailMessage
from django.conf import settings

def send_exam_report_email(subject_stats, wrong_questions, total_marks_obtained, total_marks):
    # Create in-memory Excel file
    output = io.BytesIO()
    workbook = Workbook()

    # Summary Sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_headers = ["Subject", "Attempted", "Not Attempted", "Correct", "Wrong", "Marks Obtained", "Total Marks"]
    summary_sheet.append(summary_headers)
    for subject, data in subject_stats.items():
        summary_sheet.append([
            subject,
            data['attempted'],
            data['not_attempted'],
            data['correct'],
            data['wrong'],
            data['marks_obtained'],
            data['total_marks']
        ])

    # Wrong Questions Sheet
    wrong_sheet = workbook.create_sheet(title="Wrongly Attempted Questions")
    wrong_headers = ["Question", "Your Answer", "Correct Answer"]
    wrong_sheet.append(wrong_headers)
    for wq in wrong_questions:
        wrong_sheet.append([wq['question'], wq['student_answer'], wq['correct_answer']])

    # Total Marks Sheet
    total_sheet = workbook.create_sheet(title="Total Marks")
    total_sheet.append(["Total Marks Obtained", "Total Marks"])
    total_sheet.append([total_marks_obtained, total_marks])

    workbook.save(output)
    output.seek(0)

    # Prepare Email
    email = EmailMessage(
        subject="Your Exam Report",
        body="Please find your exam report attached.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=["tilluvissu@gmail.com"],
    )

    email.attach(
        "exam_report.xlsx",
        output.read(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    email.send()
